from tkinter import *
from tkinter import filedialog
import sys
import openpyxl
import re
import os
from tkinter import messagebox
import tkinter as tk

# path = ("C:\\Users\\hi\\Documents\\Sample.xlsx")


def filebrowser():
    filename = filedialog.askopenfilename(
        initialdir="/", title="select a file", filetypes=(("Microsoft Excel Worksheet",
                                                           "*.xlsx*"), ("all files" "*.*")))

    label_file_explorer.configure(text=" file opened:" + filename)

    wb_obj = openpyxl.load_workbook(filename.strip())

    path = os.path.dirname(sys.executable)

    sheet_obj = wb_obj.active
   # window.destroy()
   # print(" Max row ", sheet_obj.max_column)
    messagebox.showinfo(
        title="Max row before", message=sheet_obj.max_column)

# sheet_obj.insert_cols(idx=1,)

    sheet_obj.insert_cols(idx=3)

    sheet_obj.insert_cols(idx=5)

    sheet_obj.cell(row=1, column=1).value = 'TEST'

    print("Max row after", sheet_obj.max_column)
    if messagebox.showinfo(
            title="Max row afert", message=sheet_obj.max_column):
        window.destroy()
# path = './Sampl.xlsx '
    path = filename
# sheet_obj.save(path)
    wb_obj.save(path)


window = tk.Tk()
# window.withdraw()
window.title("Open file")
window.geometry("500x500")
window.config(background="white")

label_file_explorer = Label(
    window, text="file", width=100, height=4, fg="blue")


button_explore = Button(window, text="browse file", command=filebrowser)


button_exit = Button(window, text="Exit", command=exit)

label_file_explorer.grid(column=1, row=1)

button_explore.grid(column=1, row=2)

button_exit.grid(column=1, row=3)

window.mainloop()


#filename = input("File Name :")
# path = "r" + path
# print(filename)
